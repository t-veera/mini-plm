"""Read a test-protocol spreadsheet as if it were a markdown table.

A protocol workbook is a table with a title block bolted on top: a few rows of preamble,
then a header row, then the tests. Once the header row is located the rows are ordinary
table rows, so they are yielded as the very same `Line`/`TableRow` records `iter_lines()`
produces for markdown. Everything downstream -- ID declaration, title/status columns,
"traces to" references -- is then the code that already runs on markdown, unchanged.

Only the sheet holding the protocol is read. A workbook's other sheets are an overview,
a run summary, a findings log; they carry no IDs and reading them would only invent
nodes out of prose.
"""
import logging
import re

from .extract import Line, TableRow

logger = logging.getLogger('files')

# The sheet that holds the tests. Checked in order, so an explicitly named "Test
# Protocol" tab wins over a generic one.
SHEET_NAME_PATTERNS = (
    re.compile(r'test\s*protocol', re.I),
    re.compile(r'protocol', re.I),
    re.compile(r'\btests?\b', re.I),
)

# How far down to hunt for the header row before giving up on a sheet. The real
# workbooks put it at row 3-5, under a title and a hardware note.
MAX_HEADER_ROW = 12
# Guard against a runaway sheet: no protocol has this many rows.
MAX_ROWS = 5000

_ID_HEADER_RE = re.compile(r'^\s*(id|tag|ref|test\s*id|item)\s*$', re.I)


def iter_sheet_lines(stream):
    """Yield Line records for the protocol sheet in `stream`, or nothing.

    `stream` is anything openpyxl can load (a file object or path). Returns an empty
    iterator -- never raises -- when the workbook cannot be opened, holds no sheet that
    looks like a protocol, or that sheet has no ID column.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - openpyxl is a hard requirement
        logger.error("traceability: openpyxl is not installed; spreadsheets cannot index")
        return

    try:
        workbook = openpyxl.load_workbook(stream, data_only=True, read_only=True)
    except Exception:
        logger.warning("traceability: could not open workbook", exc_info=True)
        return

    try:
        found = _find_protocol_sheet(workbook)
        if found is None:
            logger.warning("traceability: no sheet with an ID column in workbook (sheets: %s)",
                           ', '.join(workbook.sheetnames))
            return
        worksheet, header_row, headers = found
        for line in _rows_after(worksheet, header_row, headers):
            yield line
    finally:
        workbook.close()


def _find_protocol_sheet(workbook):
    """(worksheet, header_row_index, headers) for the sheet holding the tests, or None.

    Preferred sheet names are tried first; if none of them has a usable header the whole
    workbook is swept, so a protocol on an oddly named tab still indexes.
    """
    names = list(workbook.sheetnames)
    ordered = []
    for pattern in SHEET_NAME_PATTERNS:
        ordered.extend(name for name in names if pattern.search(name) and name not in ordered)
    ordered.extend(name for name in names if name not in ordered)

    for name in ordered:
        worksheet = workbook[name]
        header = _find_header_row(worksheet)
        if header is not None:
            return (worksheet,) + header
    return None


def _find_header_row(worksheet):
    """(row_index, headers) for the first row that carries an ID column, or None."""
    for index, row in enumerate(worksheet.iter_rows(max_row=MAX_HEADER_ROW, values_only=True), 1):
        cells = [_text(value) for value in row]
        if any(_ID_HEADER_RE.match(cell) for cell in cells):
            # Lowercased to match what iter_lines() hands the column matchers, and
            # newline-flattened so a wrapped header like "Run 1\nStatus" still reads.
            return index, [' '.join(cell.split()).lower() for cell in cells]
    return None


def _rows_after(worksheet, header_row, headers):
    for index, row in enumerate(worksheet.iter_rows(values_only=True), 1):
        if index <= header_row:
            continue
        if index - header_row > MAX_ROWS:
            logger.warning("traceability: sheet %r exceeds %s rows; stopping",
                           worksheet.title, MAX_ROWS)
            return
        cells = [_text(value) for value in row]
        if not any(cells):
            continue
        # `heading` is the sheet name: it is the only thing standing in for the document
        # heading a markdown row would inherit, and it is what a title-less row falls
        # back to.
        yield Line(index, ' '.join(cell for cell in cells if cell).strip(),
                   worksheet.title, TableRow(headers, cells))


def _text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()
